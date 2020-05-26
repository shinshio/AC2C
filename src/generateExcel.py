# coding; utf-8
'''
Generate Excel Report File.
Trim data for Excel Files.

TODO:
    No item.
'''
# import from std
import csv
# import from pypi
import datetime
import openpyxl


# set style
STYLE_FONT_PASS = openpyxl.styles.Font(
    name='Courier New',
    size=12,
    color='000000',
)
STYLE_FONT_FAIL = openpyxl.styles.Font(
    name='Courier New',
    size=12,
    color='ff0000',
)
STYLE_FILL_HEADER = openpyxl.styles.PatternFill(
    patternType='solid',
    fgColor='a6a6a6',
)
STYLE_FILL_EVEN_ROW = openpyxl.styles.PatternFill(
    patternType='solid',
    fgColor='f2f2f2',
)
STYLE_BOADER = openpyxl.styles.Border(
    left=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    right=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    top=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    bottom=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    )
)
BORDER = openpyxl.styles.Border(
    left=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    right=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    top=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    ),
    bottom=openpyxl.styles.Side(
        border_style="thin",
        color="000000"
    )
)
class Report(object):
    '''Make Excel report and level csv file
    '''
    def __init__(self, filename: str):
        self._filename = filename
        self._sample_info = []
        self._checker_info = []
        self._dmm_info = []
        self._resistor_info = []
        self._result_info = []
        self._scenario_info = []
        self._level_csv_list = []

    def __del__(self):
        pass

    def setSampleInfo(self, data: list):
        """Trim sample information of scenario to report
        Args
        ----------
            data: list
                0: event name, 1: S/N
        Returns
        ----------
            Nothing: set sample information
        """
        repSampleInfo = [
            ['イベント', data[0]],
            ['シリアル', data[1]],
        ]
        self._sample_info = repSampleInfo

    def setCheckerInfo(self, data: list):
        """Trim checker information of scenario to report
        Args
        ----------
            data: list
                0: mcu soft version, 1: AC2C version (ACCVER), 2: scenario title
        Returns
        ----------
            Nothing: set checker information
        """
        repCheckerInfo = [
            ['コントローラバージョン', data[0]],
            ['チェッカバージョン', data[1]],
            ['シナリオファイル', data[2]],
        ]
        self._checker_info = repCheckerInfo

    def setDmmInfo(self, data: list):
        """Trim dmm information of scenario to report
        Args
        ----------
            data: list
                0: maker name, 1: model name, 2: s/n
        Returns
        ----------
            Nothing: set dmm information
        """
        repDmmInfo = [
            ['メーカー', data[0]],
            ['型式', data[1]],
            ['S/N', data[2]],
        ]
        self._dmm_info = repDmmInfo

    def setResistorInfo(self, data: list):
        """Trimresistor information of scenario to report
        Args
        ----------
            data: list
                0: maker name, 1: model name, 2: s/n
        Returns
        ----------
            Nothing: set resitor information
        """
        repResistorInfo = [
            ['メーカー', data[0]],
            ['型式', data[1]],
            ['S/N', data[2]],
        ]
        self._resistor_info = repResistorInfo

    def _setResultInfo(self, data: list):
        """Trim result information to report
        Args
        ----------
            data: list
                0: checker function (SQ, SAT, LOAD)
                1: checker terminal (DAB, FSR, ...)
                2: checker status (OPEN, SHORT, ...)
                3: criteria (9011, 80011A, ...)
                4: result (9022, 80011B, ...)
                5: checker judgement (PASS, FAIL)
        Returns
        ----------
            Nothing: set result information
        """
        repResultInfo = [['機能','端子','状態','判定基準','結果','判定']]
        bufResultInfo = [''] * 6
        _data = [[d[1].split('_')[0],d[1].split('->')[0].split('_')[1],d[1].split('->')[1],d[2],d[3],d[4]] for d in data]
        resultInfo = [[s.replace(',','\n') for s in ss] for ss in _data]
        for i, item in enumerate(resultInfo):
            if item[2][0:5] == 'LEVEL':
                bufResultInfo = item
                repResultInfo.append(bufResultInfo)
                bufResultInfo = [''] * 6
            elif item[0]=='SQ' or item[0]=='SAT' or item[0]=='LOAD':
                bufResultInfo[0] = item[0]
                bufResultInfo[1] = item[1]
                bufResultInfo[2] = item[2]
            elif item[2] == 'DTCread':
                bufResultInfo[3] = item[3]
                bufResultInfo[4] = item[4]
                bufResultInfo[5] = item[5]
                if bufResultInfo[0] != '':
                    repResultInfo.append(bufResultInfo)
                    bufResultInfo = [''] * 6
        self._result_info = repResultInfo

    def setScenarioInfo(self, data: list):
        """Trim scenario information to report
        Args
        ----------
            data: list
                0: scenario number
                1: scenario order
                2: scenario criteria
                3: result
                4: judgement
        Returns
        ----------
            Nothing: set scenario information / set result information
        """
        repScenarioInfo = [['No','内容','判定基準','結果','判定']]
        _data = [[s.replace(',','\n') for s in ss] for ss in data]
        repScenarioInfo.extend(_data)
        self._scenario_info = repScenarioInfo
        self._setResultInfo(data)

    def _write2excel(self, sheet: object, data: list, start_row: int, start_col: int):
        """Write data at excel from list
        Args
        ----------
            sheet: object
                openpyxl's workbook[sheetname]
            data: list
                demension 1: rows of excel
                demension 2: columns of excel
            start_row: initial wrote cell's row of excel
            start_col: initial wrote cell's column of excel
        Returns
        ----------
            Nothing: write data into excel
        """
        for r in range(0,len(data)):
            for c in range(0,len(data[0])):
                sheet.cell(r+start_row,c+start_col).value=data[r][c]

    def outputExcelReport(self):
        """Generate Excel Report File
        Args
        ----------
            Nothing
        Returns
        ----------
            Nothing: generate excel file
        """
        # ++++++++++
        # init
        # ++++++++++
        wb = openpyxl.Workbook()
        wb.fonts = openpyxl.styles.Font(
            name = 'Courier New',
            size = 12
        )
        # create and delete sheets
        _ = wb.create_sheet(title='Cover',index=0)
        _ = wb.create_sheet(title='Results',index=1)
        _ = wb.create_sheet(title='AllItems',index=2)
        _ = wb.remove(wb.worksheets[-1])
        # ++++++++++
        # Sheet 1 <Cover>
        # ++++++++++
        ws = wb['Cover']
        # --- title and date
        timeNow = datetime.datetime.now().isoformat().split('T')[0]
        ws.merge_cells('A1:B1')
        ws.merge_cells('A3:B3')
        ws['A1'] = '納入チェック ダイアグ確認結果'
        ws['A3'] = '作成日：{}'.format(timeNow)
        # --- sample info
        ws['A5'] = '＜サンプル情報＞'
        self._write2excel(ws, self._sample_info, 6, 1)
        for r in range(6,8):
            for c in range(1,3):
                ws.cell(r,c).border = BORDER
        # --- checker info
        ws['A9'] = '＜チェッカ情報＞'
        self._write2excel(ws, self._checker_info, 10, 1)
        for r in range(10,13):
            for c in range(1,3):
                ws.cell(r,c).border = BORDER
        # --- dmm info
        ws['A14'] = '＜DMM情報＞'
        self._write2excel(ws, self._dmm_info, 15, 1)
        for r in range(15,18):
            for c in range(1,3):
                ws.cell(r,c).border = BORDER
        # --- resistor info
        ws['A19'] = '＜抵抗器情報＞'
        self._write2excel(ws, self._resistor_info, 20, 1)
        for r in range(20,23):
            for c in range(1,3):
                ws.cell(r,c).border = BORDER
        # --- set styles
        for row in ws:
            for cell in row:
                ws[cell.coordinate].font = STYLE_FONT_PASS
        # --- set column width
        for col in ws.columns:
            # init
            max_length = 0
            column = openpyxl.utils.get_column_letter(col[0].column)
            # loop
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value)) * (STYLE_FONT_PASS.size+1)/11
            # output
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        # ++++++++++
        # Sheet 2 <Results>
        # ++++++++++
        ws = wb['Results']
        # --- output all scenario
        ws['A1'] = '＜結果一覧＞'
        ws.merge_cells('A1:B1')
        self._write2excel(ws, self._result_info, 2, 1)
        for r in range(2,ws.max_row+1):
            for c in range(1,ws.max_column+1):
                ws.cell(r,c).border = BORDER
        # --- set styles
        for row in ws:
            for cell in row:
                # font color
                ws[cell.coordinate].font = STYLE_FONT_PASS
                cell.alignment = openpyxl.styles.Alignment(vertical='top')
                if cell.column==6:
                    if ws[cell.coordinate].value =='FAIL':
                        ws.cell(cell.row,1).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,2).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,3).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,4).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,5).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,6).font = STYLE_FONT_FAIL
                # cell color by header/even row
                if cell.row==2:
                    ws[cell.coordinate].fill = STYLE_FILL_HEADER
                elif cell.row%2==0:
                    ws[cell.coordinate].fill = STYLE_FILL_EVEN_ROW
                # indent in cell
                if '\n' in str(cell.value):
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        # --- set column width
        for col in ws.columns:
            # init
            max_length = 0
            column = openpyxl.utils.get_column_letter(col[0].column)
            # loop
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value)) * (STYLE_FONT_PASS.size+1)/11
            # output
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        # ++++++++++
        # Sheet 3 <AllItems>
        # ++++++++++
        ws = wb['AllItems']
        # --- output all scenario
        ws['A1'] = '＜出力一覧＞'
        ws.merge_cells('A1:B1')
        self._write2excel(ws, self._scenario_info, 2, 1)
        for r in range(2,ws.max_row+1):
            for c in range(1,ws.max_column+1):
                ws.cell(r,c).border = BORDER
        # --- set styles
        for row in ws:
            for cell in row:
                # font color
                ws[cell.coordinate].font = STYLE_FONT_PASS
                cell.alignment = openpyxl.styles.Alignment(vertical='top')
                if cell.column==5:
                    if ws[cell.coordinate].value =='FAIL':
                        ws.cell(cell.row,1).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,2).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,3).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,4).font = STYLE_FONT_FAIL
                        ws.cell(cell.row,5).font = STYLE_FONT_FAIL
                # cell color by header/even row
                if cell.row==2:
                    ws[cell.coordinate].fill = STYLE_FILL_HEADER
                elif cell.row%2==0:
                    ws[cell.coordinate].fill = STYLE_FILL_EVEN_ROW
                # indent in cell
                if '\n' in str(cell.value):
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        # --- set column width
        for col in ws.columns:
            # init
            max_length = 0
            column = openpyxl.utils.get_column_letter(col[0].column)
            # loop
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value)) * (STYLE_FONT_PASS.size+1)/11
            # output
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        # ++++++++++
        # save book
        # ++++++++++
        wb.save(self._filename)

    def outputLevelCsv(self):
        """Generate CSV File of leveling
        Args
        ----------
            Nothing
        Returns
        ----------
            Nothing: generate csv file
        """
        # extract level information from result info
        extract_level = []
        extract_level = [item for item in self._result_info if self._result_info[2][0:5]=='LEVEL']
        if extract_level == []:
            print('No Result of LEVEL')
            return None
        # copy need information
        for i, item in enumerate(extract_level):
            self._level_csv_list[i][0] = item[1]
            self._level_csv_list[i][1] = item[2].split('-')[1]
            self._level_csv_list[i][2] = item[2].split('-')[2]
            self._level_csv_list[i][3] = item[4]
        # set csv file name
        csv_file_name = self._filename.rsplit('.', 1)[1] + '.csv'
        # write csv
        with open(csv_file_name, 'w') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(self._level_csv_list)


'''
----------
main
----------
'''
if __name__ =='__main__':
    pass
